from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from pyquery import PyQuery as pq
import random
import re

def get_firm_list(filedir):
    wb_1 = load_workbook(filedir)
    ws_1 = wb_1.active
    rows = ws_1.max_row
    firm_name = []
    for i in range(4, rows+1):
        cellvalue = ws_1.cell(row=i, column=10).value
        firm_name.append(cellvalue)
    return firm_name

def write_to_excel(firm_name, patent_data, excel_name):
    wb_2 = load_workbook(excel_name)
    ws_2 = wb_2.active
    # ws_2.append([ ])
    # ws_2.append(['年份', '上市企业', patent_types[0], patent_types[1], patent_types[2]])
    for data in patent_data:
        ws_2.append([data[0], firm_name, data[1], data[2], data[3]])
    wb_2.save(excel_name)

def get_random_wait(start, end, save):    
    generate_random_number = random.uniform(start, end)
    random_wait = round(generate_random_number, save)
    return random_wait

def get_index_Formula(i,j,n):
    index_Formula = '(公开（公告）日=%s AND 申请（专利权）人=(%s)) AND (发明类型=("%s") AND 公开国家/地区/组织=(CN))'%(i,j,n)
    return index_Formula

def log_in(url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--disable-gpu')
    browser = webdriver.Chrome(executable_path=r"C:\Users\User\AppData\Local\Programs\Python\Python37-32\chromedriver.exe", chrome_options=chrome_options)
    browser.get(url)
    browser.find_element_by_id("j_username").send_keys(username)
    browser.find_element_by_id("j_password_show").send_keys(password)
    browser.find_element_by_id("wee_remember_me").click()

    time.sleep(5) #填写验证码
    browser.find_element_by_css_selector("a.btn.btn-login").click()

    WebDriverWait(browser,10).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div[3]/div/div/ul/li[2]/a')))
    browser.find_element_by_xpath("/html/body/div[1]/div[3]/div/div/ul/li[2]/a").click()
    time.sleep(2)
    return browser

def data_exist_judge(browser,firm):
    browser.execute_script("clearSearchExp();")
    test_formula = '申请（专利权）人=(%s)'%firm
    browser.find_element_by_css_selector("#searchExpDisplay").send_keys(test_formula)
    time.sleep(get_random_wait(1, 2, 2))
    browser.execute_script("excuteTableSearch();")
    WebDriverWait(browser,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="page_top"]/div/div/div')))
    # browser.find_element_by_css_selector("#tool_sort > a").click()
    page_sort_1 = browser.find_element_by_css_selector("#tool_sort > a > span")
    browser.execute_script("arguments[0].click();", page_sort_1)
    time.sleep(get_random_wait(0.2, 1, 2))
    page_sort_2 = browser.find_element_by_css_selector("#tool_sort > a > ul > li:nth-child(4) > a")
    browser.execute_script("arguments[0].click();", page_sort_2)
    time.sleep(get_random_wait(0.5, 1, 2))
    html = browser.page_source
    patent_info_all = re.findall(re_rule_1, html, re.S)
    if patent_info_all:
        pass
    else:
        content = browser.find_element_by_css_selector("div.page_top")
        detail_content = content.text
        patent_info_all = re.findall(re_rule_2, detail_content, re.S) 
        patent_info_all = [one.strip() for one in patent_info_all]

    begin_date = browser.find_element_by_css_selector("#resultMode > div > div.list-container > ul > li:nth-child(1) > div > div.item-content.clear > div.item-content-body.left > p:nth-child(4) > a").text
    begin_date = begin_date.split(".")
    print(begin_date)
    slices_date = date[0:date.index(int(begin_date[0]))]
    print(slices_date)
    print(patent_info_all)
    label = 0
    if int(patent_info_all[0]) <= 0:
        return browser, label, slices_date, begin_date
    else:
        label = 1
        return browser, label, slices_date, begin_date

def page_spider(firm,browser,begin_date):
    patent_data = []
    print('>> %s <<———————>> 有 <<详细专利数据！'%firm)
    for year in date[date.index(int(begin_date[0])):]:
        print('>> 正在 <<爬取第 %s 个企业——第 %s 年数据！'%(firm_name_list.index(firm), year))
        patent_data_year = []
        patent_data_year.append(year)
        count = 0
        for patent_type in types:
            try:
                browser.find_element_by_css_selector("body > div.ui-popup.ui-popup-modal.ui-popup-show.ui-popup-focus")
                browser.find_element_by_css_selector("body > div.ui-popup.ui-popup-modal.ui-popup-show.ui-popup-focus > div > table > tbody > tr:nth-child(3) > td > div.ui-dialog-button > button").click()
                time.sleep(get_random_wait(0.5, 1.2, 2))
                print("clear the alert!")
            except:
                pass

            print('>> 正在 <<爬取第 %s 个企业——第 %s 年数据！'%(firm_name_list.index(firm), year))
            search_formula = get_index_Formula(year, firm, patent_type)
            print(search_formula)
            browser.execute_script("clearSearchExp();")
            # browser.find_element_by_css_selector("a.btn.btn-remove").click()
            browser.find_element_by_css_selector("#searchExpDisplay").send_keys(search_formula)
            time.sleep(get_random_wait(3, 4.5, 2))
            # browser.find_element_by_css_selector("a.btn.btn-search").click()
            browser.execute_script("excuteTableSearch();")

            WebDriverWait(browser,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="page_top"]/div/div/div')))
            # browser.execute_script("document.querySelector("#resultlistForm_top")")
            time.sleep(get_random_wait(2, 3, 2))
            html = browser.page_source
            # print(html)
            patent_info = re.findall(re_rule_1, html, re.S)
            if patent_info:
                pass
            else:
                content = browser.find_element_by_css_selector("div.page_top")
                detail_content = content.text
                patent_info = re.findall(re_rule_2, detail_content, re.S) 
                patent_info = [one.strip() for one in patent_info]
                
            print(patent_info)

            if int(patent_info[0]) == count:
                time.sleep(get_random_wait(1, 2, 2))
            elif int(patent_info[0]) < 0:
                patent_info.pop()
                patent_info.append('0')
            else:
                pass
            count = int(patent_info[0])
            patent_data_year.append(count)
        patent_data.append(patent_data_year)

    return patent_data


if __name__ == '__main__':
    path_r = r'C:\Users\User\Desktop\上市公司.xlsx'
    path_w = r'C:\Users\User\Desktop\上市公司专利情况.xlsx'
    url = 'http://pss-system.cnipa.gov.cn/sipopublicsearch/portal/uilogin-forwardLogin.shtml'

    username = 'aha_heihei'
    password = 'qw260480092'
    re_rule_1 = r'共.*?页&nbsp;(.*?)&nbsp;条数据'
    re_rule_2 = r'共.*?页(.*?)条数据'

    date = [year for year in range(2000,2019)]
    types = ['I','U','D']
    patent_types = ['发明专利','实用新型','外观设计']

    firm_name_list = get_firm_list(path_r)
    browser = log_in(url)
    for firm in firm_name_list[551:800]:
        print('-----------------------------------------------------------------------------------------------------------')
        print('>> 开始 <<爬取第 %s 个企业！'%firm_name_list.index(firm))
        browser, label, slices_date, begin_date = data_exist_judge(browser,firm)
        if label == 0:
            patent_data = [[year, 0, 0, 0] for year in date]
            write_to_excel(firm, patent_data, path_w)
            print('>> %s <<———————>> 无 <<专利数据！'%firm)
        elif label == 1:
            print('>> %s <<———————>> 有 <<详细专利数据！'%firm)
            patent_data_null = [[year, 0, 0, 0] for year in slices_date]
            write_to_excel(firm, patent_data_null, path_w)
            patent_data = page_spider(firm, browser, begin_date)

        write_to_excel(firm, patent_data, path_w)





# for firm in firm_name_list[547:800]:
#     print('-----------------------------------------------------------------------------------------------------------')
#     print('>> 开始 <<爬取第 %s 个企业！'%firm_name_list.index(firm))
#     patent_data = []
#     # trick = 0
#     browser.execute_script("clearSearchExp();")
#     test_formula = '申请（专利权）人=(%s)'%firm
#     browser.find_element_by_css_selector("#searchExpDisplay").send_keys(test_formula)
#     time.sleep(get_random_wait(1, 2, 2))
#     browser.execute_script("excuteTableSearch();")
#     WebDriverWait(browser,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="page_top"]/div/div/div')))
#     # browser.find_element_by_css_selector("#tool_sort > a").click()
#     page_sort_1 = browser.find_element_by_css_selector("#tool_sort > a > span")
#     browser.execute_script("arguments[0].click();", page_sort_1)
#     time.sleep(get_random_wait(0.2, 1, 2))
#     page_sort_2 = browser.find_element_by_css_selector("#tool_sort > a > ul > li:nth-child(4) > a")
#     browser.execute_script("arguments[0].click();", page_sort_2)
#     time.sleep(get_random_wait(0.5, 1, 2))
#     html = browser.page_source
#     patent_info_all = re.findall(re_rule_1, html, re.S)
#     if patent_info_all:
#         pass
#     else:
#         content = browser.find_element_by_css_selector("div.page_top")
#         detail_content = content.text
#         patent_info_all = re.findall(re_rule_2, detail_content, re.S) 
#         patent_info_all = [one.strip() for one in patent_info_all]

#     begin_date = browser.find_element_by_css_selector("#resultMode > div > div.list-container > ul > li:nth-child(1) > div > div.item-content.clear > div.item-content-body.left > p:nth-child(4) > a").text
#     begin_date = begin_date.split(".")
#     print(begin_date)
#     slices_date = date[0:date.index(int(begin_date[0]))]
#     print(slices_date)
#     print(patent_info_all)

#     if int(patent_info_all[0]) <= 0:
#         patent_data = [[year, 0, 0, 0] for year in date]
#         write_to_excel(firm, patent_data, path_w)
#         print('>> %s <<———————>> 无 <<专利数据！'%firm)
#     else:
#         print('>> %s <<———————>> 有 <<详细专利数据！'%firm)
#         for year in date[date.index(int(begin_date[0])):]:
#             print('>> 正在 <<爬取第 %s 个企业——第 %s 年数据！'%(firm_name_list.index(firm), year))
#             patent_data_year = []
#             patent_data_year.append(year)
#             count = 0
#             for patent_type in types:
#                 try:
#                     browser.find_element_by_css_selector("body > div.ui-popup.ui-popup-modal.ui-popup-show.ui-popup-focus")
#                     browser.find_element_by_css_selector("body > div.ui-popup.ui-popup-modal.ui-popup-show.ui-popup-focus > div > table > tbody > tr:nth-child(3) > td > div.ui-dialog-button > button").click()
#                     time.sleep(get_random_wait(0.5, 1.2, 2))
#                     print("clear the alert!")
#                 except:
#                     pass

#                 print('>> 正在 <<爬取第 %s 个企业——第 %s 年数据！'%(firm_name_list.index(firm), year))
#                 search_formula = get_index_Formula(year, firm, patent_type)
#                 print(search_formula)
#                 browser.execute_script("clearSearchExp();")
#                 # browser.find_element_by_css_selector("a.btn.btn-remove").click()
#                 browser.find_element_by_css_selector("#searchExpDisplay").send_keys(search_formula)
#                 time.sleep(get_random_wait(3, 4.5, 2))
#                 # browser.find_element_by_css_selector("a.btn.btn-search").click()
#                 browser.execute_script("excuteTableSearch();")

#                 WebDriverWait(browser,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="page_top"]/div/div/div')))
#                 # browser.execute_script("document.querySelector("#resultlistForm_top")")
#                 time.sleep(get_random_wait(2, 3, 2))
#                 html = browser.page_source
#                 # print(html)
#                 patent_info = re.findall(re_rule_1, html, re.S)
#                 if patent_info:
#                     pass
#                 else:
#                     content = browser.find_element_by_css_selector("div.page_top")
#                     detail_content = content.text
#                     patent_info = re.findall(re_rule_2, detail_content, re.S) 
#                     patent_info = [one.strip() for one in patent_info]
                    
#                 print(patent_info)

#                 if int(patent_info[0]) == count:
#                     time.sleep(get_random_wait(1, 2, 2))
#                 elif int(patent_info[0]) < 0:
#                     patent_info.pop()
#                     patent_info.append('0')
#                 else:
#                     pass
#                 count = int(patent_info[0])
#                 patent_data_year.append(count)
#             patent_data.append(patent_data_year)
#         write_to_excel(firm, patent_data, path_w)